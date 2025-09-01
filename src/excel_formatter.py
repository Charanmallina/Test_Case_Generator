# FILE: src/excel_formatter.py

import json
import os
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl for Excel support...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

class ExcelFormatter:
    """Convert test cases to Excel format for QA teams"""
    
    def __init__(self):
        print("📊 Excel Formatter initialized")
    
    def convert_to_excel(self, json_file_path: str, excel_file_path: str) -> bool:
        """
        Convert JSON test cases to Excel format
        
        Args:
            json_file_path: Path to generated test cases JSON
            excel_file_path: Path to save Excel file
        """
        try:
            # Load test cases
            with open(json_file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            
            test_cases = data.get('test_cases', [])
            metadata = data.get('metadata', {})
            
            if not test_cases:
                print("No test cases found in JSON file")
                return False
            
            print(f"Converting {len(test_cases)} test cases to Excel...")
            
            # Create workbook
            wb = Workbook()
            
            # Create worksheets
            summary_ws = wb.active
            summary_ws.title = "Summary"
            
            test_cases_ws = wb.create_sheet("Test Cases")
            details_ws = wb.create_sheet("Detailed Steps")
            
            # Format summary sheet
            self._create_summary_sheet(summary_ws, test_cases, metadata)
            
            # Format test cases sheet
            self._create_test_cases_sheet(test_cases_ws, test_cases)
            
            # Format detailed steps sheet
            self._create_details_sheet(details_ws, test_cases)
            
            # Save the workbook
            wb.save(excel_file_path)
            print(f"✅ Excel file saved: {excel_file_path}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error converting to Excel: {str(e)}")
            return False
    
    def _create_summary_sheet(self, ws, test_cases, metadata):
        """Create summary overview sheet"""
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws['A1'] = "QA Test Cases Generation Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        ws[f'A{row}'] = "Generated At:"
        ws[f'B{row}'] = metadata.get('generated_at', '')
        row += 1
        
        ws[f'A{row}'] = "Total Test Cases:"
        ws[f'B{row}'] = len(test_cases)
        row += 1
        
        ws[f'A{row}'] = "Success Rate:"
        ws[f'B{row}'] = metadata.get('success_rate', '')
        row += 1
        
        ws[f'A{row}'] = "AI Model Used:"
        ws[f'B{row}'] = metadata.get('model_used', '')
        row += 2
        
        # Priority breakdown
        ws[f'A{row}'] = "Priority Breakdown"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        priorities = {}
        for tc in test_cases:
            priority = tc.get('priority', 'Unknown')
            priorities[priority] = priorities.get(priority, 0) + 1
        
        for priority, count in priorities.items():
            ws[f'A{row}'] = priority
            ws[f'B{row}'] = count
            row += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_test_cases_sheet(self, ws, test_cases):
        """Create main test cases sheet"""
        
        # Headers
        headers = [
            "Test Case ID", "Domain", "Service", "Test Type", "Priority", 
            "Severity", "Issue Description", "Test Scenario", 
            "Expected Result", "Actual Issue", "Automation Feasibility", 
            "Customer Impact", "Source Call ID"
        ]
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write test case data
        for row, tc in enumerate(test_cases, 2):
            ws.cell(row=row, column=1, value=tc.get('test_case_id', ''))
            ws.cell(row=row, column=2, value=tc.get('domain', ''))
            ws.cell(row=row, column=3, value=tc.get('service', ''))
            ws.cell(row=row, column=4, value=tc.get('test_type', ''))
            ws.cell(row=row, column=5, value=tc.get('priority', ''))
            ws.cell(row=row, column=6, value=tc.get('severity', ''))
            ws.cell(row=row, column=7, value=tc.get('issue_description', ''))
            ws.cell(row=row, column=8, value=tc.get('test_scenario', ''))
            ws.cell(row=row, column=9, value=tc.get('expected_result', ''))
            ws.cell(row=row, column=10, value=tc.get('actual_issue', ''))
            ws.cell(row=row, column=11, value=tc.get('automation_feasibility', ''))
            ws.cell(row=row, column=12, value=tc.get('customer_impact', ''))
            ws.cell(row=row, column=13, value=tc.get('source_call_id', ''))
            
            # Color code by priority
            priority = tc.get('priority', '').lower()
            if priority == 'critical':
                fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif priority == 'high':
                fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
            elif priority == 'medium':
                fill = PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid")
            else:
                fill = None
            
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = fill
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_details_sheet(self, ws, test_cases):
        """Create detailed test steps sheet"""
        
        # Headers
        headers = ["Test Case ID", "Step Number", "Action", "Expected Result"]
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write detailed steps
        row = 2
        for tc in test_cases:
            test_case_id = tc.get('test_case_id', '')
            test_steps = tc.get('test_steps', [])
            
            if isinstance(test_steps, list):
                for step_num, step in enumerate(test_steps, 1):
                    ws.cell(row=row, column=1, value=test_case_id)
                    ws.cell(row=row, column=2, value=step_num)
                    ws.cell(row=row, column=3, value=step)
                    ws.cell(row=row, column=4, value="")  # QA can fill expected results
                    row += 1
            else:
                # Handle case where test_steps is a string
                ws.cell(row=row, column=1, value=test_case_id)
                ws.cell(row=row, column=2, value=1)
                ws.cell(row=row, column=3, value=str(test_steps))
                ws.cell(row=row, column=4, value="")
                row += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width


def convert_json_to_excel(json_file: str, excel_file: str):
    """Convert existing JSON test cases to Excel format"""
    
    formatter = ExcelFormatter()
    success = formatter.convert_to_excel(json_file, excel_file)
    
    if success:
        print(f"✅ Conversion completed!")
        print(f"📁 Excel file: {excel_file}")
        print(f"📊 Open in Excel to view formatted test cases")
    else:
        print("❌ Conversion failed")


def test_excel_conversion():
    """Test converting your generated test cases to Excel"""
    
    # Check for existing JSON test cases
    json_files = []
    output_dir = 'data/output'
    
    if os.path.exists(output_dir):
        files = os.listdir(output_dir)
        json_files = [f for f in files if f.endswith('.json') and 'test_cases_' in f]
    
    if not json_files:
        print("❌ No test case JSON files found in data/output/")
        print("Generate test cases first using the web app or direct AI generator")
        return
    
    # Use the most recent file
    latest_file = max(json_files)
    json_path = os.path.join(output_dir, latest_file)
    
    # Create Excel filename
    excel_filename = latest_file.replace('.json', '.xlsx')
    excel_path = os.path.join(output_dir, excel_filename)
    
    print(f"Converting: {json_path}")
    print(f"To Excel: {excel_path}")
    
    convert_json_to_excel(json_path, excel_path)


if __name__ == "__main__":
    test_excel_conversion()